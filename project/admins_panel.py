# admins_panel.py
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QLineEdit, QComboBox,
    QFrame, QMessageBox, QDialog, QFormLayout, QStackedWidget, QDateEdit
)
from PyQt6.QtCore import Qt, QDate
import pymysql
from datetime import datetime


def safe_item(value):
    return QTableWidgetItem("" if value is None else str(value))


class UserDialog(QDialog):
    def __init__(self, parent=None, user=None):
        super().__init__(parent)
        self.setWindowTitle("User Form")
        self.setFixedWidth(400)

        layout = QVBoxLayout(self)
        form = QFormLayout()
        form.setSpacing(12)

        self.username_input = QLineEdit()
        self.password_input = QLineEdit()
        self.confirm_password_input = QLineEdit()
        self.role_input = QComboBox()
        self.role_input.addItems(["admin", "cashier"])
        self.role_input.setStyleSheet("""
        QComboBox { color: black; background: white; }

        QComboBox QAbstractItemView {
                color: black;
                background-color: white;
                selection-background-color: #007BFF;
                selection-color: white;
            }
        """)

        form.addRow("Username:", self.username_input)
        form.addRow("Password:", self.password_input)
        form.addRow("Confirm Password:", self.confirm_password_input)
        form.addRow("Role:", self.role_input)
        layout.addLayout(form)

        btns = QHBoxLayout()
        self.save_btn = QPushButton("Save")
        self.cancel_btn = QPushButton("Cancel")
        btns.addWidget(self.save_btn)
        btns.addWidget(self.cancel_btn)
        layout.addLayout(btns)

        self.cancel_btn.clicked.connect(self.reject)
        self.save_btn.clicked.connect(self.accept)

        if user:
            self.username_input.setText(user["username"])
            self.role_input.setCurrentText(user["role"])

    def get_data(self):
        return {
            "username": self.username_input.text().strip(),
            "password": self.password_input.text().strip(),
            "confirm_password": self.confirm_password_input.text().strip(),
            "role": self.role_input.currentText()
        }


class AdminsPanel(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Admin Tools")
        self.selected_row = None

        main_layout = QVBoxLayout(self)

        # ===== Header =====
        title = QLabel("Admin Tools")
        subtitle = QLabel("Manage users, inventory, transactions, and reports")
        title.setStyleSheet("font-size: 22px; font-weight: bold; margin-top: 15px;")
        subtitle.setStyleSheet("font-size: 13px; color: gray; margin-bottom: 15px;")
        main_layout.addWidget(title)
        main_layout.addWidget(subtitle)

        # ===== Switch Buttons =====
        switch_layout = QHBoxLayout()
        self.user_btn = QPushButton("User & Role Management")
        self.inv_btn = QPushButton("Inventory & Stock Control")
        self.trans_btn = QPushButton("Transaction History")
        self.reports_btn = QPushButton("Generate Reports")

        for btn in (self.user_btn, self.inv_btn, self.trans_btn, self.reports_btn):
            btn.setStyleSheet("""
                QPushButton {
                    background: #f5f5f5;
                    border: 1px solid #ccc;
                    border-radius: 6px;
                    padding: 8px 16px;
                    font-weight: bold;
                    color: black;
                }
                QPushButton:hover {
                    background: #007BFF;
                    color: white;
                }
                QPushButton:checked {
                    background: #007BFF;
                    color: white;
                }
            """)
            btn.setCheckable(True)
            switch_layout.addWidget(btn)

        main_layout.addLayout(switch_layout)

        # ===== Stacked Widget =====
        self.stack = QStackedWidget()
        self.user_card = self.create_user_management_card()
        self.inv_card = self.create_inventory_card()
        self.trans_card = self.create_transaction_history_card()
        self.reports_card = self.create_reports_card()

        self.stack.addWidget(self.user_card)
        self.stack.addWidget(self.inv_card)
        self.stack.addWidget(self.trans_card)
        self.stack.addWidget(self.reports_card)
        main_layout.addWidget(self.stack)

        # Default view
        self.user_btn.setChecked(True)
        self.stack.setCurrentWidget(self.user_card)

        # Button switching
        self.user_btn.clicked.connect(lambda: self.switch_view(self.user_card, self.user_btn))
        self.inv_btn.clicked.connect(lambda: self.switch_view(self.inv_card, self.inv_btn))
        self.trans_btn.clicked.connect(lambda: self.switch_view(self.trans_card, self.trans_btn))
        self.reports_btn.clicked.connect(lambda: self.switch_view(self.reports_card, self.reports_btn))

        # Load data
        self.load_users()
        self.load_inventory()

    def switch_view(self, widget, btn):
        self.stack.setCurrentWidget(widget)
        self.user_btn.setChecked(False)
        self.inv_btn.setChecked(False)
        self.trans_btn.setChecked(False)
        self.reports_btn.setChecked(False)
        btn.setChecked(True)

        # Load data when switching to transaction history
        if widget == self.trans_card:
            self.load_transactions()

    # ===== DB =====
    def db_connect(self):
        return pymysql.connect(
            host="localhost", user="root", password="",
            database="techstore_pos", cursorclass=pymysql.cursors.DictCursor
        )

    # ===== User Management =====
    def create_user_management_card(self):
        card = QFrame()
        card.setStyleSheet("QFrame { background: white; border-radius: 10px; border: 1px solid #ddd; padding: 12px; }")
        layout = QVBoxLayout(card)

        controls = QHBoxLayout()
        self.search_user = QLineEdit()
        self.search_user.setPlaceholderText("Search users...")
        controls.addWidget(self.search_user, stretch=2)

        self.add_btn = QPushButton("Add User")
        self.edit_btn = QPushButton("Edit User")
        self.delete_btn = QPushButton("Delete User")
        for btn in (self.add_btn, self.edit_btn, self.delete_btn):
            btn.setStyleSheet("background:#007BFF;color:white;padding:6px 12px;border-radius:6px;")
            controls.addWidget(btn)
        layout.addLayout(controls)

        self.users_table = QTableWidget()
        self.users_table.setColumnCount(3)
        self.users_table.setHorizontalHeaderLabels(["ID", "Username", "Role"])
        self.style_table(self.users_table)
        layout.addWidget(self.users_table)

        self.users_table.cellClicked.connect(self.on_row_click)
        self.add_btn.clicked.connect(self.add_user)
        self.edit_btn.clicked.connect(self.edit_user)
        self.delete_btn.clicked.connect(self.delete_user)

        return card

    def load_users(self):
        try:
            conn = self.db_connect()
            cur = conn.cursor()
            cur.execute("SELECT id, username, role FROM users")
            rows = cur.fetchall()
            conn.close()
        except Exception as e:
            print("⚠️ Error loading users:", e)
            rows = []

        self.users_table.setRowCount(0)
        for row in rows:
            r = self.users_table.rowCount()
            self.users_table.insertRow(r)
            self.users_table.setItem(r, 0, safe_item(row.get("id")))
            self.users_table.setItem(r, 1, safe_item(row.get("username")))
            self.users_table.setItem(r, 2, safe_item(row.get("role")))

    def on_row_click(self, row, col):
        self.selected_row = row

    def add_user(self):
        dialog = UserDialog(self)
        if dialog.exec():
            data = dialog.get_data()
            if data["password"] != data["confirm_password"]:
                QMessageBox.warning(self, "Error", "Passwords do not match!")
                return
            try:
                conn = self.db_connect()
                cur = conn.cursor()
                cur.execute("INSERT INTO users (username,password,role) VALUES (%s,%s,%s)",
                            (data["username"], data["password"], data["role"]))
                conn.commit();
                conn.close()
            except Exception as e:
                QMessageBox.critical(self, "DB Error", str(e))
            self.load_users()

    def edit_user(self):
        if self.selected_row is None: return
        uid = int(self.users_table.item(self.selected_row, 0).text())
        user = {"username": self.users_table.item(self.selected_row, 1).text(),
                "role": self.users_table.item(self.selected_row, 2).text()}
        dialog = UserDialog(self, user)
        if dialog.exec():
            data = dialog.get_data()
            if data["password"] and data["password"] != data["confirm_password"]:
                QMessageBox.warning(self, "Error", "Passwords do not match!");
                return
            try:
                conn = self.db_connect();
                cur = conn.cursor()
                if data["password"]:
                    cur.execute("UPDATE users SET username=%s,password=%s,role=%s WHERE id=%s",
                                (data["username"], data["password"], data["role"], uid))
                else:
                    cur.execute("UPDATE users SET username=%s,role=%s WHERE id=%s",
                                (data["username"], data["role"], uid))
                conn.commit();
                conn.close()
            except Exception as e:
                QMessageBox.critical(self, "DB Error", str(e))
            self.load_users()

    def delete_user(self):
        if self.selected_row is None: return
        uid = int(self.users_table.item(self.selected_row, 0).text())
        uname = self.users_table.item(self.selected_row, 1).text()
        confirm = QMessageBox.question(self, "Confirm Delete",
                                       f"Delete user '{uname}'?",
                                       QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if confirm == QMessageBox.StandardButton.Yes:
            try:
                conn = self.db_connect();
                cur = conn.cursor()
                cur.execute("DELETE FROM users WHERE id=%s", (uid,))
                conn.commit();
                conn.close()
            except Exception as e:
                QMessageBox.critical(self, "DB Error", str(e))
            self.load_users()

    # ===== Inventory =====
    def create_inventory_card(self):
        card = QFrame()
        card.setStyleSheet("QFrame { background: white; border-radius: 10px; border: 1px solid #ddd; padding: 12px; }")
        layout = QVBoxLayout(card)

        self.inventory_table = QTableWidget()
        self.inventory_table.setColumnCount(3)
        self.inventory_table.setHorizontalHeaderLabels(["ID", "Product", "Stock"])
        self.style_table(self.inventory_table)
        layout.addWidget(self.inventory_table)

        return card

    def load_inventory(self):
        try:
            conn = self.db_connect()
            cur = conn.cursor()
            cur.execute("SELECT id, name, stock FROM products ORDER BY id ASC")
            rows = cur.fetchall()
            conn.close()
        except Exception as e:
            print("⚠️ Error loading inventory:", e)
            rows = []

        self.inventory_table.setRowCount(0)
        for row in rows:
            r = self.inventory_table.rowCount()
            self.inventory_table.insertRow(r)
            self.inventory_table.setItem(r, 0, safe_item(row.get("id")))
            self.inventory_table.setItem(r, 1, safe_item(row.get("name")))
            self.inventory_table.setItem(r, 2, safe_item(row.get("stock")))

    # ===== Transaction History =====
    def create_transaction_history_card(self):
        card = QFrame()
        card.setStyleSheet("QFrame { background: white; border-radius: 10px; border: 1px solid #ddd; padding: 12px; }")
        layout = QVBoxLayout(card)

        # Filters
        filter_layout = QHBoxLayout()
        filter_label = QLabel("Filter by Month:")
        filter_layout.addWidget(filter_label)

        self.month_combo = QComboBox()
        months = ["January", "February", "March", "April", "May", "June",
                  "July", "August", "September", "October", "November", "December"]
        self.month_combo.addItems(months)
        self.month_combo.setCurrentIndex(QDate.currentDate().month() - 1)
        self.month_combo.setMinimumWidth(150)
        self.month_combo.setStyleSheet("""
            QComboBox { 
                color: black; 
                background: white; 
                padding: 5px 10px; 
                min-width: 150px;
            }
            QComboBox QAbstractItemView {
                color: black;
                background-color: white;
                selection-background-color: #007BFF;
                selection-color: white;
            }
        """)
        filter_layout.addWidget(self.month_combo)

        self.year_combo = QComboBox()
        current_year = QDate.currentDate().year()
        for year in range(current_year - 5, current_year + 1):
            self.year_combo.addItem(str(year))
        self.year_combo.setCurrentText(str(current_year))
        self.year_combo.setMinimumWidth(100)
        self.year_combo.setStyleSheet("""
            QComboBox { 
                color: black; 
                background: white; 
                padding: 5px 10px; 
                min-width: 100px;
            }
            QComboBox QAbstractItemView {
                color: black;
                background-color: white;
                selection-background-color: #007BFF;
                selection-color: white;
            }
        """)
        filter_layout.addWidget(self.year_combo)

        self.filter_btn = QPushButton("Apply Filter")
        self.filter_btn.setStyleSheet("background:#007BFF;color:white;padding:6px 12px;border-radius:6px;")
        self.filter_btn.clicked.connect(self.load_transactions)
        filter_layout.addWidget(self.filter_btn)

        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        # Transaction table - 4 columns (removed Cashier)
        self.transaction_table = QTableWidget()
        self.transaction_table.setColumnCount(4)
        self.transaction_table.setHorizontalHeaderLabels(
            ["Transaction ID", "Date", "Time", "Total Amount"])
        self.transaction_table.horizontalHeader().setVisible(True)
        self.style_table(self.transaction_table)
        layout.addWidget(self.transaction_table)

        # Summary
        summary_layout = QHBoxLayout()
        self.trans_summary = QLabel("Total Transactions: 0 | Total Revenue: ₱0.00")
        self.trans_summary.setStyleSheet("font-weight: bold; font-size: 14px; margin-top: 10px;")
        summary_layout.addWidget(self.trans_summary)
        summary_layout.addStretch()
        layout.addLayout(summary_layout)

        return card

    def load_transactions(self):
        self.transaction_table.setRowCount(0)
        total_revenue = 0
        row_count = 0

        try:
            conn = self.db_connect()
            cur = conn.cursor()

            month = self.month_combo.currentIndex() + 1
            year = int(self.year_combo.currentText())

            # Only select columns that exist: id, created_at, total
            query = """
                SELECT id, created_at, total
                FROM transactions
                WHERE MONTH(created_at) = %s AND YEAR(created_at) = %s
                ORDER BY created_at DESC
            """
            cur.execute(query, (month, year))
            rows = cur.fetchall()
            conn.close()

            for row in rows:
                r = self.transaction_table.rowCount()
                self.transaction_table.insertRow(r)

                # Parse datetime
                trans_date = row.get("created_at")
                date_str = ""
                time_str = ""

                if trans_date:
                    if isinstance(trans_date, str):
                        try:
                            trans_date = datetime.strptime(trans_date, "%Y-%m-%d %H:%M:%S")
                        except:
                            pass
                    if hasattr(trans_date, 'strftime'):
                        date_str = trans_date.strftime("%Y-%m-%d")
                        time_str = trans_date.strftime("%H:%M:%S")

                # Only 4 columns: ID, Date, Time, Total Amount (Cashier removed)
                self.transaction_table.setItem(r, 0, safe_item(row.get("id")))
                self.transaction_table.setItem(r, 1, safe_item(date_str))
                self.transaction_table.setItem(r, 2, safe_item(time_str))
                self.transaction_table.setItem(r, 3, safe_item(f"₱{row.get('total', 0):.2f}"))

                total_revenue += float(row.get("total", 0))
                row_count += 1

        except Exception as e:
            print("⚠️ Error loading transactions:", e)
            QMessageBox.critical(self, "Database Error", f"Error loading transactions: {str(e)}")

        month_name = self.month_combo.currentText()
        self.trans_summary.setText(
            f"{month_name} {year}: {row_count} Transactions | Total Revenue: ₱{total_revenue:,.2f}")

    # ===== Reports =====
    def create_reports_card(self):
        card = QFrame()
        card.setStyleSheet("QFrame { background: white; border-radius: 10px; border: 1px solid #ddd; padding: 12px; }")
        layout = QVBoxLayout(card)

        # Report header
        header = QLabel("Generate Sales & Inventory Reports")
        header.setStyleSheet("font-size: 16px; font-weight: bold; margin-bottom: 20px;")
        layout.addWidget(header)

        # Report options
        report_grid = QHBoxLayout()

        # Sales report section
        sales_section = QVBoxLayout()
        sales_title = QLabel("Sales Reports")
        sales_title.setStyleSheet("font-weight: bold; font-size: 14px;")
        sales_section.addWidget(sales_title)

        self.daily_sales_btn = QPushButton("Daily Sales Report")
        self.monthly_sales_btn = QPushButton("Monthly Sales Report")
        self.yearly_sales_btn = QPushButton("Yearly Sales Report")

        for btn in (self.daily_sales_btn, self.monthly_sales_btn, self.yearly_sales_btn):
            btn.setStyleSheet("""
                QPushButton {
                    background:#28a745;
                    color:white;
                    padding:10px;
                    border-radius:6px;
                    margin:5px;
                    border: none;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background:#218838;
                }
                QPushButton:pressed {
                    background:#1e7e34;
                }
            """)
            sales_section.addWidget(btn)

        report_grid.addLayout(sales_section)

        # Inventory report section
        inv_section = QVBoxLayout()
        inv_title = QLabel("Inventory Reports")
        inv_title.setStyleSheet("font-weight: bold; font-size: 14px;")
        inv_section.addWidget(inv_title)

        self.low_stock_btn = QPushButton("Low Stock Report")
        self.stock_summary_btn = QPushButton("Stock Summary Report")
        self.product_sales_btn = QPushButton("Product Sales Report")

        for btn in (self.low_stock_btn, self.stock_summary_btn, self.product_sales_btn):
            btn.setStyleSheet("""
                QPushButton {
                    background:#17a2b8;
                    color:white;
                    padding:10px;
                    border-radius:6px;
                    margin:5px;
                    border: none;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background:#138496;
                }
                QPushButton:pressed {
                    background:#0c5460;
                }
            """)
            inv_section.addWidget(btn)

        report_grid.addLayout(inv_section)
        layout.addLayout(report_grid)

        # Report display area
        layout.addSpacing(20)
        report_label = QLabel("Report Preview")
        report_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        layout.addWidget(report_label)

        self.report_table = QTableWidget()
        self.style_table(self.report_table)
        layout.addWidget(self.report_table)

        # Export buttons
        export_layout = QHBoxLayout()
        export_layout.addStretch()
        self.export_csv_btn = QPushButton("Export to Excel")
        self.export_csv_btn.setStyleSheet("""
            QPushButton {
                background:#007BFF;
                color:white;
                padding:8px 16px;
                border-radius:6px;
                border: none;
                font-weight: bold;
            }
            QPushButton:hover {
                background:#0056b3;
            }
            QPushButton:pressed {
                background:#004085;
            }
        """)
        export_layout.addWidget(self.export_csv_btn)
        layout.addLayout(export_layout)

        # Connect buttons
        self.daily_sales_btn.clicked.connect(self.generate_daily_sales)
        self.monthly_sales_btn.clicked.connect(self.generate_monthly_sales)
        self.yearly_sales_btn.clicked.connect(self.generate_yearly_sales)
        self.low_stock_btn.clicked.connect(self.generate_low_stock)
        self.stock_summary_btn.clicked.connect(self.generate_stock_summary)
        self.product_sales_btn.clicked.connect(self.generate_product_sales)
        self.export_csv_btn.clicked.connect(self.export_report)

        return card

    def generate_daily_sales(self):
        try:
            conn = self.db_connect()
            cur = conn.cursor()
            cur.execute("""
                SELECT DATE(created_at) as date, COUNT(*) as transactions, SUM(total) as total
                FROM transactions
                WHERE DATE(created_at) = CURDATE()
                GROUP BY DATE(created_at)
            """)
            rows = cur.fetchall()
            conn.close()

            if not rows:
                rows = [{"date": QDate.currentDate().toString("yyyy-MM-dd"), "transactions": 0, "total": 0.00}]

            self.display_report(["Date", "Transactions", "Total Sales"], rows)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to generate report: {str(e)}")

    def generate_monthly_sales(self):
        try:
            conn = self.db_connect()
            cur = conn.cursor()
            cur.execute("""
                SELECT DATE_FORMAT(created_at, '%Y-%m') as month, 
                       COUNT(*) as transactions, 
                       SUM(total) as total
                FROM transactions
                GROUP BY DATE_FORMAT(created_at, '%Y-%m')
                ORDER BY month DESC
                LIMIT 12
            """)
            rows = cur.fetchall()
            conn.close()

            if not rows:
                rows = [{"month": QDate.currentDate().toString("yyyy-MM"), "transactions": 0, "total": 0.00}]

            self.display_report(["Month", "Transactions", "Total Sales"], rows)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to generate report: {str(e)}")

    def generate_yearly_sales(self):
        try:
            conn = self.db_connect()
            cur = conn.cursor()
            cur.execute("""
                SELECT YEAR(created_at) as year, 
                       COUNT(*) as transactions, 
                       SUM(total) as total
                FROM transactions
                GROUP BY YEAR(created_at)
                ORDER BY year DESC
            """)
            rows = cur.fetchall()
            conn.close()

            if not rows:
                rows = [{"year": QDate.currentDate().year(), "transactions": 0, "total": 0.00}]

            self.display_report(["Year", "Transactions", "Total Sales"], rows)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to generate report: {str(e)}")

    def generate_low_stock(self):
        try:
            conn = self.db_connect()
            cur = conn.cursor()
            cur.execute("""
                SELECT id, name, stock, price
                FROM products
                WHERE stock < 10
                ORDER BY stock ASC
            """)
            rows = cur.fetchall()
            conn.close()

            self.display_report(["Product ID", "Name", "Stock", "Price"], rows)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to generate report: {str(e)}")

    def generate_stock_summary(self):
        try:
            conn = self.db_connect()
            cur = conn.cursor()
            cur.execute("""
                SELECT id, name, stock, price, (stock * price) as value
                FROM products
                ORDER BY value DESC
            """)
            rows = cur.fetchall()
            conn.close()

            self.display_report(["Product ID", "Name", "Stock", "Price", "Total Value"], rows)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to generate report: {str(e)}")

    def generate_product_sales(self):
        try:
            conn = self.db_connect()
            cur = conn.cursor()
            cur.execute("""
                SELECT p.name as name, 
                       SUM(ti.quantity) as total_sold,
                       SUM(ti.quantity * ti.price) as revenue
                FROM transaction_items ti
                JOIN products p ON ti.product_id = p.id
                GROUP BY p.id, p.name
                ORDER BY revenue DESC
            """)
            rows = cur.fetchall()
            conn.close()

            if not rows:
                QMessageBox.information(self, "No Data", "No product sales data available.")
                return

            self.display_report(["Product", "Units Sold", "Revenue"], rows)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to generate report: {str(e)}")

    def display_report(self, headers, data):
        self.report_table.setColumnCount(len(headers))
        self.report_table.setHorizontalHeaderLabels(headers)
        self.report_table.setRowCount(0)

        for row_data in data:
            r = self.report_table.rowCount()
            self.report_table.insertRow(r)
            for col, key in enumerate(row_data.keys()):
                value = row_data[key]
                if isinstance(value, float):
                    value = f"₱{value:,.2f}" if "total" in key.lower() or "revenue" in key.lower() or "value" in key.lower() or "price" in key.lower() else f"{value:,.2f}"
                self.report_table.setItem(r, col, safe_item(value))

    def export_report(self):
        if self.report_table.rowCount() == 0:
            QMessageBox.warning(self, "No Data", "Generate a report first before exporting.")
            return

        from PyQt6.QtWidgets import QFileDialog
        filename, _ = QFileDialog.getSaveFileName(self, "Export Report", "", "Excel Files (*.xlsx)")

        if filename:
            try:
                # Using openpyxl for Excel export
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment

                wb = Workbook()
                ws = wb.active
                ws.title = "Report"

                # Write headers with styling
                headers = []
                for col in range(self.report_table.columnCount()):
                    headers.append(self.report_table.horizontalHeaderItem(col).text())
                ws.append(headers)

                # Style header row
                header_fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Write data
                for row in range(self.report_table.rowCount()):
                    row_data = []
                    for col in range(self.report_table.columnCount()):
                        item = self.report_table.item(row, col)
                        row_data.append(item.text() if item else "")
                    ws.append(row_data)

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

                wb.save(filename)
                QMessageBox.information(self, "Success", f"Report exported to {filename}")
            except ImportError:
                QMessageBox.critical(self, "Missing Library",
                                     "openpyxl library is required for Excel export.\n"
                                     "Install it with: pip install openpyxl")
            except Exception as e:
                QMessageBox.critical(self, "Export Error", str(e))

    # ===== Table Styling =====
    def style_table(self, table):
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        table.verticalHeader().setVisible(False)
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        table.setAlternatingRowColors(True)

        # Set explicit colors for table text and background
        table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                color: black;
                gridline-color: #ddd;
                border: 1px solid #ddd;
            }
            QTableWidget::item {
                color: black;
                padding: 8px;
            }
            QTableWidget::item:selected {
                background-color: #007BFF;
                color: white;
            }
            QTableWidget::item:alternate {
                background-color: #f9f9f9;
            }
            QHeaderView::section {
                background-color: #f1f1f1;
                color: #000000;
                padding: 8px;
                border: 1px solid #ccc;
                font-weight: bold;
                font-size: 14px;
                height: 50px;
            }
        """)